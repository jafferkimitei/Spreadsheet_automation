from pathlib import Path
import os

import openpyxl
import pytest


env_task_root = os.environ.get("TASK_ROOT_FOR_TESTS")
if env_task_root:
    TASK_ROOT = Path(env_task_root).resolve()
else:
    TASK_ROOT = Path.cwd().resolve()

WORKBOOK_PATH = TASK_ROOT / "output" / "finance_report.xlsx"

EXPECTED_SHEETS = {
    "Raw Cleaned Data",
    "Regional Summary",
    "Product Summary",
    "Risk Metrics",
    "Validation Checks",
}


@pytest.fixture(scope="module")
def workbook():
    assert WORKBOOK_PATH.exists(), "Expected output/finance_report.xlsx to be created"
    return openpyxl.load_workbook(WORKBOOK_PATH, data_only=False)


def sheet_records(ws):
    rows = list(ws.iter_rows(values_only=True))
    assert rows, f"{ws.title} should not be empty"
    headers = [str(value).strip() for value in rows[0]]
    records = []
    for row in rows[1:]:
        if all(value is None for value in row):
            continue
        records.append(dict(zip(headers, row)))
    return records


def get_record(records, key, expected_value):
    for record in records:
        if record.get(key) == expected_value:
            return record
    raise AssertionError(f"Could not find row where {key} == {expected_value!r}")


def assert_close(actual, expected, tolerance=1e-6):
    assert actual is not None, f"Expected numeric value {expected}, got None"
    assert abs(float(actual) - expected) <= tolerance, (
        f"Expected {expected}, got {actual}"
    )


def test_workbook_file_is_created():
    assert WORKBOOK_PATH.exists(), "The finance workbook was not generated"


def test_required_sheets_exist(workbook):
    assert EXPECTED_SHEETS.issubset(set(workbook.sheetnames))


def test_raw_cleaned_data_excludes_reversed_transactions(workbook):
    records = sheet_records(workbook["Raw Cleaned Data"])

    transaction_ids = {record["transaction_id"] for record in records}
    statuses = {record["status"] for record in records}

    assert len(records) == 9
    assert "reversed" not in statuses
    assert {"TX-1003", "TX-1008", "TX-1011"}.isdisjoint(transaction_ids)


def test_raw_cleaned_data_contains_region_and_kes_columns(workbook):
    records = sheet_records(workbook["Raw Cleaned Data"])
    first = records[0]

    assert "world_bank_region" in first
    assert "revenue_kes" in first
    assert "cost_kes" in first

    tx_1001 = get_record(records, "transaction_id", "TX-1001")
    assert tx_1001["world_bank_region"] == "Sub-Saharan Africa"
    assert_close(tx_1001["revenue_kes"], 12500 * 129.50)
    assert_close(tx_1001["cost_kes"], 7100 * 129.50)


def test_regional_summary_totals_are_correct(workbook):
    records = sheet_records(workbook["Regional Summary"])

    ssa = get_record(records, "world_bank_region", "Sub-Saharan Africa")
    south_asia = get_record(records, "world_bank_region", "South Asia")
    lac = get_record(records, "world_bank_region", "Latin America & Caribbean")

    assert_close(ssa["revenue_usd"], 50900)
    assert_close(ssa["cost_usd"], 27000)
    assert_close(ssa["gross_profit_usd"], 23900)
    assert_close(ssa["margin_pct"], 23900 / 50900)

    assert_close(south_asia["revenue_usd"], 31700)
    assert_close(south_asia["cost_usd"], 18100)

    assert_close(lac["revenue_usd"], 34600)
    assert_close(lac["gross_profit_usd"], 16000)


def test_product_summary_margin_percentage_is_correct(workbook):
    records = sheet_records(workbook["Product Summary"])

    loans = get_record(records, "product", "Loans")
    insurance = get_record(records, "product", "Insurance")
    payments = get_record(records, "product", "Payments")

    assert_close(loans["revenue_usd"], 26800)
    assert_close(loans["margin_pct"], 11600 / 26800)

    assert_close(insurance["revenue_usd"], 43500)
    assert_close(insurance["margin_pct"], 20800 / 43500)

    assert_close(payments["revenue_usd"], 46900)
    assert_close(payments["margin_pct"], 21100 / 46900)


def test_risk_metrics_ignore_missing_values(workbook):
    records = sheet_records(workbook["Risk Metrics"])

    ssa = get_record(records, "world_bank_region", "Sub-Saharan Africa")
    south_asia = get_record(records, "world_bank_region", "South Asia")
    lac = get_record(records, "world_bank_region", "Latin America & Caribbean")

    assert_close(ssa["avg_credit_risk_ratio"], 0.089)
    assert_close(ssa["avg_restructured_portfolio_ratio"], 0.03375)

    assert_close(south_asia["avg_credit_risk_ratio"], 0.07)
    assert_close(south_asia["avg_restructured_portfolio_ratio"], 0.031)

    assert_close(lac["avg_credit_risk_ratio"], 0.0645)
    assert_close(lac["avg_restructured_portfolio_ratio"], 0.0215)


def test_validation_sheet_reports_core_control_totals(workbook):
    records = sheet_records(workbook["Validation Checks"])
    checks = {record["check_name"]: record["value"] for record in records}

    assert checks["source_row_count"] == 12
    assert checks["clean_row_count"] == 9
    assert checks["excluded_reversed_count"] == 3
    assert_close(checks["approved_revenue_usd_total"], 117200)
    assert_close(checks["approved_revenue_kes_total"], 117200 * 129.50)


def test_header_formatting_is_applied_to_all_sheets(workbook):
    for sheet_name in EXPECTED_SHEETS:
        ws = workbook[sheet_name]
        header_cells = list(ws[1])
        assert header_cells, f"{sheet_name} should have a header row"

        for cell in header_cells:
            assert cell.font.bold, f"{sheet_name} header {cell.coordinate} should be bold"
            assert cell.fill.fill_type == "solid", (
                f"{sheet_name} header {cell.coordinate} should have a solid fill"
            )
            assert cell.alignment.horizontal == "center", (
                f"{sheet_name} header {cell.coordinate} should be center-aligned"
            )


def test_region_and_product_transaction_counts_reconcile_to_clean_rows(workbook):
    raw_records = sheet_records(workbook["Raw Cleaned Data"])
    region_records = sheet_records(workbook["Regional Summary"])
    product_records = sheet_records(workbook["Product Summary"])

    clean_count = len(raw_records)
    region_count_total = sum(int(record["transaction_count"]) for record in region_records)
    product_count_total = sum(int(record["transaction_count"]) for record in product_records)

    assert region_count_total == clean_count
    assert product_count_total == clean_count


def test_summary_revenue_totals_match_validation_checks(workbook):
    validation_records = sheet_records(workbook["Validation Checks"])
    validation = {record["check_name"]: float(record["value"]) for record in validation_records}

    region_records = sheet_records(workbook["Regional Summary"])
    product_records = sheet_records(workbook["Product Summary"])

    region_revenue_total = sum(float(record["revenue_usd"]) for record in region_records)
    product_revenue_total = sum(float(record["revenue_usd"]) for record in product_records)

    assert_close(region_revenue_total, validation["approved_revenue_usd_total"])
    assert_close(product_revenue_total, validation["approved_revenue_usd_total"])


def test_raw_cleaned_data_transaction_ids_are_unique(workbook):
    records = sheet_records(workbook["Raw Cleaned Data"])
    transaction_ids = [record["transaction_id"] for record in records]
    assert len(transaction_ids) == len(set(transaction_ids)), (
        "Raw Cleaned Data should not contain duplicate transaction_id values"
    )
